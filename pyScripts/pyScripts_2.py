from openpyxl import Workbook

wb = Workbook()
# creating excel sheet
ws1 = wb.create_sheet("MySheet")
# changing sheet name
ws1.title = "MySheet-1"
print(wb.sheetnames)